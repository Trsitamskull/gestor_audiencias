from openpyxl import load_workbook
from datetime import datetime
from typing import List, Tuple, Optional
from .audiencia import Audiencia


class ExcelManager:
    """Gestiona las operaciones con archivos Excel."""

    FILA_INICIO_DATOS = 11
    FILA_MAXIMA_DATOS_PARA_LIMPIAR = 110
    FILA_TOTALES = 111

    def __init__(self, archivo_path: str):
        self.archivo_path = archivo_path

    def guardar_audiencia(self, audiencia: Audiencia) -> bool:
        """Guarda una nueva audiencia en el archivo Excel."""
        try:
            wb = load_workbook(self.archivo_path)
            ws = wb.active

            if ws is None:
                raise Exception("No se pudo acceder a la hoja de trabajo")

            # Encontrar la primera fila vacía
            fila_destino = self.FILA_INICIO_DATOS
            while ws.cell(row=fila_destino, column=2).value is not None:
                fila_destino += 1

            # Escribir los datos
            self._escribir_fila_audiencia(ws, fila_destino, audiencia)

            wb.save(self.archivo_path)
            return True

        except PermissionError:
            raise Exception(
                "No se pudo guardar. Cierre el archivo de Excel si está abierto."
            )
        except Exception as e:
            raise Exception(f"Error al guardar: {e}")

    def actualizar_audiencia(self, fila: int, audiencia: Audiencia) -> bool:
        """Actualiza una audiencia existente."""
        try:
            wb = load_workbook(self.archivo_path)
            ws = wb.active

            if ws is None:
                raise Exception("No se pudo acceder a la hoja de trabajo")

            self._escribir_fila_audiencia(ws, fila, audiencia)
            wb.save(self.archivo_path)
            return True

        except Exception as e:
            raise Exception(f"Error al actualizar: {e}")

    def _escribir_fila_audiencia(self, ws, fila: int, audiencia: Audiencia):
        """Escribe los datos de una audiencia en una fila específica."""
        ws.cell(row=fila, column=2, value=audiencia.radicado)
        ws.cell(row=fila, column=3, value=audiencia.tipo)
        ws.cell(row=fila, column=4, value=audiencia.fecha)
        ws.cell(row=fila, column=5, value=audiencia.hora)
        ws.cell(row=fila, column=6, value=audiencia.juzgado)
        ws.cell(row=fila, column=7, value=audiencia.realizada_si)
        ws.cell(row=fila, column=8, value=audiencia.realizada_no)

        for i, motivo in enumerate(audiencia.motivos):
            ws.cell(row=fila, column=9 + i, value=motivo)

        ws.cell(row=fila, column=17, value=audiencia.observaciones)

    def leer_registros(self) -> List[Tuple[int, List]]:
        """Lee todos los registros del archivo."""
        try:
            wb = load_workbook(self.archivo_path, read_only=True)
            ws = wb.active

            if ws is None:
                raise Exception("No se pudo acceder a la hoja de trabajo")

            registros = []
            for fila_num in range(self.FILA_INICIO_DATOS, ws.max_row + 1):
                datos_fila = [cell.value for cell in ws[fila_num]]
                if datos_fila[1]:  # Si hay radicado
                    registros.append((fila_num, datos_fila))

            return registros

        except Exception as e:
            raise Exception(f"Error al leer registros: {e}")

    def reordenar_y_guardar(self) -> Tuple[int, int, List[int]]:
        """Reordena todos los registros por fecha/hora y calcula totales."""
        try:
            wb = load_workbook(self.archivo_path)
            ws = wb.active

            if ws is None:
                raise Exception("No se pudo acceder a la hoja de trabajo")

            # Leer y ordenar datos
            datos_tabla = []
            for row in ws.iter_rows(
                min_row=self.FILA_INICIO_DATOS, max_row=ws.max_row, values_only=True
            ):
                if not row or not row[1]:  # Si no hay radicado
                    continue
                try:
                    f = datetime.strptime(str(row[3]), "%d/%m/%Y")
                    h = datetime.strptime(str(row[4]), "%H:%M")
                    datos_tabla.append((f, h, list(row)))
                except (ValueError, TypeError):
                    continue

            datos_tabla.sort(key=lambda x: (x[0], x[1]), reverse=True)

            # Limpiar área de datos
            self._limpiar_area_datos(ws)

            # Escribir datos ordenados
            for idx, (_, _, fila_datos) in enumerate(datos_tabla, start=1):
                ws.cell(row=self.FILA_INICIO_DATOS + idx - 1, column=1, value=idx)
                for col, val in enumerate(fila_datos[1:], start=2):
                    ws.cell(row=self.FILA_INICIO_DATOS + idx - 1, column=col, value=val)

            # Calcular y escribir totales
            num_registros, total_si, totales_motivos = self._calcular_totales(ws)
            self._escribir_totales(ws, total_si, totales_motivos)

            wb.save(self.archivo_path)
            return num_registros, total_si, totales_motivos

        except Exception as e:
            raise Exception(f"Error al reordenar: {e}")

    def _limpiar_area_datos(self, ws):
        """Limpia el área de datos en la hoja."""
        for i in range(self.FILA_INICIO_DATOS, self.FILA_MAXIMA_DATOS_PARA_LIMPIAR + 1):
            for j in range(1, 21):
                ws.cell(row=i, column=j, value=None)

    def _calcular_totales(self, ws) -> Tuple[int, int, List[int]]:
        """Calcula los totales de la hoja."""
        totales_motivos = [0] * 8
        total_si = 0
        num_registros = 0

        for fila in range(
            self.FILA_INICIO_DATOS, self.FILA_MAXIMA_DATOS_PARA_LIMPIAR + 1
        ):
            if ws.cell(row=fila, column=2).value:  # Si hay radicado
                num_registros += 1
                if ws.cell(row=fila, column=7).value == "SI":
                    total_si += 1
                for i in range(8):
                    if ws.cell(row=fila, column=9 + i).value:
                        totales_motivos[i] += 1

        return num_registros, total_si, totales_motivos

    def _escribir_totales(self, ws, total_si: int, totales_motivos: List[int]):
        """Escribe los totales en la hoja."""
        ws.cell(row=self.FILA_TOTALES, column=7, value=total_si)
        for i, total in enumerate(totales_motivos):
            ws.cell(row=self.FILA_TOTALES, column=9 + i, value=total)

    def contar_registros(self) -> int:
        """Cuenta el número de registros en el archivo."""
        try:
            wb = load_workbook(self.archivo_path, read_only=True)
            ws = wb.active
            num_registros, _, _ = self._calcular_totales(ws)
            return num_registros
        except Exception:
            return 0
