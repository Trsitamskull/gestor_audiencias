import tkinter as tk
from tkinter import ttk, messagebox, simpledialog
from datetime import datetime
from openpyxl import load_workbook
import os
import ttkbootstrap as tb
from gestor_archivos import (
    crear_copia_plantilla,
    listar_archivos_creados,
    seleccionar_archivo,
    eliminar_archivo,
    descargar_archivo,
)

# --- Variables Globales y Constantes ---
ARCHIVO_EXCEL = None  # No hay archivo seleccionado al inicio
FILA_INICIO_DATOS = 11  # Los datos empiezan desde esta fila
FILA_MAXIMA_DATOS_PARA_LIMPIAR = 110  # Fila máxima para limpiar datos
FILA_TOTALES = 111  # Fila para los totales
MODO_EDICION = False  # Controla si estamos en modo de edición
FILA_EDITANDO = None  # Almacena el número de la fila que se está editando
contador_registros_label = None

# === Interfaz Gráfica con tkinter y ttkbootstrap ===
ventana = tb.Window(themename="flatly")
ventana.title("Ordenador de Audiencias")
ventana.geometry("600x600")  # Ajustado para nuevos botones
ventana.minsize(600, 600)
ventana.maxsize(1200, 1080)
ventana.resizable(True, True)

try:
    ventana.state("zoomed")
except tk.TclError:
    ventana.attributes("-fullscreen", True)

# --- Configuración de Grid y Fuentes ---
for i in range(13):  # Aumentado para nuevas filas
    ventana.rowconfigure(i, weight=1)
ventana.columnconfigure(1, weight=1)
fuente = ("Segoe UI", 11)

# === Widgets de la Interfaz ===

# Contador de Registros
contador_registros_label = ttk.Label(
    ventana, text="Registros: 0", font=("Segoe UI", 10, "bold")
)
contador_registros_label.grid(row=0, column=1, sticky="ne", padx=20, pady=5)

# Campo: Radicado del proceso
tk.Label(ventana, text="Radicado del proceso:", bg="#f7f7fa", font=fuente).grid(
    row=1, column=0, sticky="e", padx=16, pady=8
)
entrada_radicado = tk.Entry(ventana, width=40, font=fuente)
entrada_radicado.grid(row=1, column=1, sticky="ew", padx=16, pady=8)

# Campo: Tipo de audiencia
tipos_audiencia = [
    "Alegatos de conclusión",
    "Audiencia concentrada",
    "Audiencia de acusación",
    "Audiencia de juicio oral",
    "Audiencia de preclusion",
    "Audiencia preliminar",
    "Audiencia preparatoria",
    "Otra",
]
tk.Label(ventana, text="Tipo de audiencia:", bg="#f7f7fa", font=fuente).grid(
    row=2, column=0, sticky="e", padx=16, pady=8
)
frame_tipo = tk.Frame(ventana, bg="#f7f7fa")
frame_tipo.grid(row=2, column=1, sticky="ew", padx=16, pady=8)
frame_tipo.columnconfigure(0, weight=1)
combo_tipo = ttk.Combobox(
    frame_tipo, values=tipos_audiencia, state="readonly", font=fuente, width=28
)
combo_tipo.grid(row=0, column=0, sticky="ew")
entrada_tipo_tra = tk.Entry(frame_tipo, width=25, font=fuente)
entrada_tipo_tra.grid(row=0, column=1, sticky="w", padx=(8, 0))
entrada_tipo_tra.grid_remove()  # Oculto por defecto

# Campo: Fecha
tk.Label(ventana, text="Fecha (DD/MM/AAAA):", bg="#f7f7fa", font=fuente).grid(
    row=3, column=0, sticky="e", padx=16, pady=8
)
frame_fecha = tk.Frame(ventana, bg="#f7f7fa")
frame_fecha.grid(row=3, column=1, sticky="w", padx=16, pady=8)
dias = [f"{i:02d}" for i in range(1, 32)]
meses = [f"{i:02d}" for i in range(1, 13)]
anios = [str(a) for a in range(datetime.now().year, datetime.now().year + 6)]
combo_dia = ttk.Combobox(frame_fecha, values=dias, width=3, font=fuente)
combo_mes = ttk.Combobox(frame_fecha, values=meses, width=3, font=fuente)
combo_anio = ttk.Combobox(frame_fecha, values=anios, width=5, font=fuente)
combo_dia.grid(row=0, column=0)
tk.Label(frame_fecha, text="/", bg="#f7f7fa", font=fuente).grid(row=0, column=1)
combo_mes.grid(row=0, column=2)
tk.Label(frame_fecha, text="/", bg="#f7f7fa", font=fuente).grid(row=0, column=3)
combo_anio.grid(row=0, column=4)

# Campo: Hora
tk.Label(ventana, text="Hora (militar):", bg="#f7f7fa", font=fuente).grid(
    row=4, column=0, sticky="e", padx=16, pady=8
)
frame_hora = tk.Frame(ventana, bg="#f7f7fa")
frame_hora.grid(row=4, column=1, sticky="w", padx=16, pady=8)
spin_hora = ttk.Spinbox(
    frame_hora, from_=0, to=23, width=3, format="%02.0f", font=fuente
)
spin_hora.grid(row=0, column=0)
tk.Label(frame_hora, text=":", bg="#f7f7fa", font=fuente).grid(row=0, column=1)
spin_minuto = ttk.Spinbox(
    frame_hora, from_=0, to=59, width=3, format="%02.0f", font=fuente
)
spin_minuto.grid(row=0, column=2)

# Campo: Juzgado
tk.Label(ventana, text="Juzgado:", bg="#f7f7fa", font=fuente).grid(
    row=5, column=0, sticky="e", padx=16, pady=8
)
entrada_juzgado = tk.Entry(ventana, width=40, font=fuente)
entrada_juzgado.grid(row=5, column=1, sticky="ew", padx=16, pady=8)

# Campo: ¿Se realizó?
tk.Label(ventana, text="¿Se realizó?", bg="#f7f7fa", font=fuente).grid(
    row=6, column=0, sticky="e", padx=16, pady=8
)
combo_realizada = ttk.Combobox(
    ventana, values=["SI", "NO"], state="readonly", font=fuente
)
combo_realizada.grid(row=6, column=1, sticky="w", padx=16, pady=8)

# Campo: Motivos de no realización
frame_motivos = tk.LabelFrame(
    ventana, text="Motivos (si NO se realizó)", font=fuente, bg="#f7f7fa"
)
frame_motivos.grid(row=7, column=0, columnspan=2, padx=16, pady=16, sticky="ew")
frame_motivos.columnconfigure((0, 1, 2, 3), weight=1)
motivo_etiquetas = [
    "Juez",
    "Fiscalía",
    "Usuario",
    "INPEC",
    "Víctima",
    "ICBF",
    "Defensor Confianza",
    "Defensor Público",
]
motivo_vars = [tk.StringVar() for _ in motivo_etiquetas]
checkboxes_motivos = []
for i, texto in enumerate(motivo_etiquetas):
    chk = tk.Checkbutton(
        frame_motivos,
        text=texto,
        variable=motivo_vars[i],
        onvalue=texto,
        offvalue="",
        bg="#f7f7fa",
        font=fuente,
    )
    chk.grid(row=i // 4, column=i % 4, sticky="w", padx=4, pady=2)
    checkboxes_motivos.append(chk)

# Campo: Observaciones
tk.Label(ventana, text="Observaciones:", bg="#f7f7fa", font=fuente).grid(
    row=8, column=0, sticky="ne", padx=16, pady=8
)
frame_obs = tk.Frame(ventana, bg="#f7f7fa")
frame_obs.grid(row=8, column=1, sticky="nsew", padx=16, pady=8)
frame_obs.rowconfigure(0, weight=1)
frame_obs.columnconfigure(0, weight=1)
entrada_observaciones = tk.Text(frame_obs, width=40, height=4, font=fuente)
entrada_observaciones.grid(row=0, column=0, sticky="nsew")
scroll_obs = ttk.Scrollbar(
    frame_obs, orient="vertical", command=entrada_observaciones.yview
)
scroll_obs.grid(row=0, column=1, sticky="ns")
entrada_observaciones.config(yscrollcommand=scroll_obs.set)

# --- Botones de Acción (Guardar, Actualizar, Cancelar) ---
frame_botones_accion = tk.Frame(ventana, bg="#f7f7fa")
frame_botones_accion.grid(row=9, column=0, columnspan=2, pady=20, padx=16, sticky="ew")
frame_botones_accion.columnconfigure((0, 1, 2), weight=1)

btn_guardar = tk.Button(
    frame_botones_accion,
    text="Guardar",
    bg="#4a90e2",
    fg="white",
    font=(fuente[0], 12, "bold"),
)
btn_guardar.grid(
    row=0, column=0, columnspan=3, sticky="ew", padx=2
)  # Ocupa todo el espacio al inicio

btn_actualizar = tk.Button(
    frame_botones_accion,
    text="Actualizar",
    bg="#28a745",
    fg="white",
    font=(fuente[0], 12, "bold"),
)
btn_actualizar.grid(row=0, column=0, sticky="ew", padx=2)
btn_actualizar.grid_remove()  # Oculto por defecto

btn_cancelar_edicion = tk.Button(
    frame_botones_accion,
    text="Cancelar",
    bg="#dc3545",
    fg="white",
    font=(fuente[0], 12, "bold"),
)
btn_cancelar_edicion.grid(row=0, column=1, sticky="ew", padx=2)
btn_cancelar_edicion.grid_remove()  # Oculto por defecto

# --- Lógica de la Aplicación (Funciones) ---


def mostrar_entrada_otra(event=None):
    """Muestra u oculta el campo de texto para 'Otra' audiencia."""
    if combo_tipo.get() == "Otra":
        entrada_tipo_tra.grid()
        entrada_tipo_tra.focus_set()
    else:
        entrada_tipo_tra.grid_remove()
        entrada_tipo_tra.delete(0, tk.END)


def habilitar_motivos(event=None):
    """Habilita o deshabilita los checkboxes de motivos."""
    estado = "normal" if combo_realizada.get() == "NO" else "disabled"
    for chk in checkboxes_motivos:
        chk.config(state=estado)
    if estado == "disabled":
        for var in motivo_vars:
            var.set("")


def limpiar_campos():
    """Limpia todos los campos del formulario."""
    entrada_radicado.delete(0, tk.END)
    combo_tipo.set("")
    entrada_tipo_tra.delete(0, tk.END)
    hoy = datetime.now()
    combo_dia.set(f"{hoy.day:02d}")
    combo_mes.set(f"{hoy.month:02d}")
    combo_anio.set(str(hoy.year))
    spin_hora.set("00")
    spin_minuto.set("00")
    entrada_juzgado.delete(0, tk.END)
    combo_realizada.set("")
    for var in motivo_vars:
        var.set("")
    entrada_observaciones.delete("1.0", tk.END)
    habilitar_motivos()
    mostrar_entrada_otra()
    entrada_radicado.focus_set()


def activar_modo_edicion():
    """Cambia la interfaz al modo de edición."""
    global MODO_EDICION
    MODO_EDICION = True
    btn_guardar.grid_remove()
    btn_actualizar.grid()
    btn_cancelar_edicion.grid()
    ventana.title("Ordenador de Audiencias - EDITANDO REGISTRO")


def desactivar_modo_edicion():
    """Vuelve la interfaz al modo normal."""
    global MODO_EDICION, FILA_EDITANDO
    MODO_EDICION = False
    FILA_EDITANDO = None
    btn_actualizar.grid_remove()
    btn_cancelar_edicion.grid_remove()
    btn_guardar.grid()
    ventana.title("Ordenador de Audiencias")
    limpiar_campos()


def cargar_datos_para_edicion(fila_datos):
    """Carga los datos de una fila en el formulario para editarlos."""
    limpiar_campos()

    # Radicado (índice 1), Tipo (2), Fecha (3), Hora (4), Juzgado (5), etc.
    entrada_radicado.insert(0, str(fila_datos[1] or ""))

    tipo = str(fila_datos[2] or "")
    if tipo in tipos_audiencia:
        combo_tipo.set(tipo)
    else:
        combo_tipo.set("Otra")
        entrada_tipo_tra.insert(0, tipo)
        mostrar_entrada_otra()

    if fila_datos[3] and isinstance(fila_datos[3], str) and "/" in fila_datos[3]:
        dia, mes, anio = fila_datos[3].split("/")
        combo_dia.set(dia)
        combo_mes.set(mes)
        combo_anio.set(anio)

    if fila_datos[4] and isinstance(fila_datos[4], str) and ":" in fila_datos[4]:
        hora, minuto = fila_datos[4].split(":")
        spin_hora.set(f"{int(hora):02d}")
        spin_minuto.set(f"{int(minuto):02d}")

    entrada_juzgado.insert(0, str(fila_datos[5] or ""))

    if fila_datos[6] == "SI":
        combo_realizada.set("SI")
    elif fila_datos[7] == "NO":
        combo_realizada.set("NO")

    for i, var in enumerate(motivo_vars):
        if len(fila_datos) > i + 8 and fila_datos[i + 8]:
            var.set(str(fila_datos[i + 8]))

    if len(fila_datos) > 16 and fila_datos[16]:
        entrada_observaciones.insert("1.0", str(fila_datos[16]))

    habilitar_motivos()


def obtener_datos_formulario():
    """Recoge y valida los datos del formulario."""
    radicado = entrada_radicado.get().strip()
    tipo = (
        entrada_tipo_tra.get().strip()
        if combo_tipo.get() == "Otra"
        else combo_tipo.get()
    )
    fecha_str = f"{combo_dia.get()}/{combo_mes.get()}/{combo_anio.get()}"
    hora_str = f"{int(spin_hora.get()):02d}:{int(spin_minuto.get()):02d}"
    juzgado = entrada_juzgado.get().strip()
    realizada = combo_realizada.get()

    if not (radicado and tipo and realizada and juzgado):
        if not messagebox.askyesno(
            "Advertencia", "Hay campos obligatorios vacíos. ¿Desea continuar?"
        ):
            return None

    try:
        dt_fecha = datetime.strptime(fecha_str, "%d/%m/%Y")
    except ValueError:
        messagebox.showerror("Error de Fecha", f"La fecha '{fecha_str}' no es válida.")
        return None

    if realizada == "NO" and not any(var.get() for var in motivo_vars):
        messagebox.showerror(
            "Error de Motivo",
            "Debe seleccionar un motivo si la audiencia NO se realizó.",
        )
        return None

    datos = {
        "radicado": radicado,
        "tipo": tipo,
        "fecha": dt_fecha.strftime("%d/%m/%Y"),
        "hora": hora_str,
        "juzgado": juzgado,
        "realizada_si": "SI" if realizada == "SI" else "",
        "realizada_no": "NO" if realizada == "NO" else "",
        "motivos": [var.get() or "" for var in motivo_vars],
        "observaciones": entrada_observaciones.get("1.0", tk.END).strip(),
    }
    return datos


def reordenar_y_guardar_excel():
    """Lee todos los registros, los reordena por fecha/hora y guarda el archivo."""
    if not ARCHIVO_EXCEL:
        return

    try:
        wb = load_workbook(ARCHIVO_EXCEL)
        ws = wb.active
        if ws is None:
            messagebox.showerror("Error", "No se pudo acceder a la hoja de trabajo")
            return False

        datos_tabla = []
        for row in ws.iter_rows(
            min_row=FILA_INICIO_DATOS, max_row=ws.max_row, values_only=True
        ):
            if not row or not row[1]:
                continue
            try:
                f = datetime.strptime(str(row[3]), "%d/%m/%Y")
                h = datetime.strptime(str(row[4]), "%H:%M")
                datos_tabla.append((f, h, list(row)))
            except (ValueError, TypeError):
                continue

        datos_tabla.sort(key=lambda x: (x[0], x[1]), reverse=True)

        # Limpiar área de datos
        for i in range(FILA_INICIO_DATOS, FILA_MAXIMA_DATOS_PARA_LIMPIAR + 1):
            for j in range(1, 21):
                ws.cell(row=i, column=j, value=None)

        # Escribir datos ordenados
        for idx, (_, _, fila_datos) in enumerate(datos_tabla, start=1):
            ws.cell(row=FILA_INICIO_DATOS + idx - 1, column=1, value=idx)  # N°
            for col, val in enumerate(fila_datos[1:], start=2):
                ws.cell(row=FILA_INICIO_DATOS + idx - 1, column=col, value=val)

        num_registros, total_si, totales_motivos = sumar_y_actualizar_totales(ws)

        # Escribir totales
        ws.cell(row=FILA_TOTALES, column=7, value=total_si)
        for i, total in enumerate(totales_motivos):
            ws.cell(row=FILA_TOTALES, column=9 + i, value=total)

        wb.save(ARCHIVO_EXCEL)
        actualizar_contador_registros()
        return True

    except PermissionError:
        messagebox.showerror(
            "Error de Permisos",
            "No se pudo guardar. Cierre el archivo de Excel si está abierto.",
        )
    except Exception as e:
        messagebox.showerror(
            "Error Inesperado", f"Ocurrió un error al procesar el archivo: {e}"
        )
    return False


def guardar_datos():
    """Guarda un nuevo registro en el archivo Excel."""
    if not ARCHIVO_EXCEL:
        messagebox.showwarning(
            "Sin Archivo", "Por favor, seleccione un archivo de trabajo."
        )
        return

    datos_nuevos = obtener_datos_formulario()
    if not datos_nuevos:
        return

    if not isinstance(ARCHIVO_EXCEL, str):
        messagebox.showerror("Error", "Ruta de archivo inválida")
        return

    archivo = str(ARCHIVO_EXCEL)
    try:
        wb = load_workbook(archivo)
        ws = wb.active
        if ws is None:
            messagebox.showerror("Error", "No se pudo acceder a la hoja de trabajo")
            return

        # Encontrar la primera fila vacía para el nuevo registro
        fila_destino = FILA_INICIO_DATOS
        while ws.cell(row=fila_destino, column=2).value is not None:
            fila_destino += 1

        # Escribir el nuevo registro
        ws.cell(row=fila_destino, column=2, value=datos_nuevos["radicado"])
        ws.cell(row=fila_destino, column=3, value=datos_nuevos["tipo"])
        ws.cell(row=fila_destino, column=4, value=datos_nuevos["fecha"])
        ws.cell(row=fila_destino, column=5, value=datos_nuevos["hora"])
        ws.cell(row=fila_destino, column=6, value=datos_nuevos["juzgado"])
        ws.cell(row=fila_destino, column=7, value=datos_nuevos["realizada_si"])
        ws.cell(row=fila_destino, column=8, value=datos_nuevos["realizada_no"])
        for i, motivo in enumerate(datos_nuevos["motivos"]):
            ws.cell(row=fila_destino, column=9 + i, value=motivo)
        ws.cell(row=fila_destino, column=17, value=datos_nuevos["observaciones"])

        wb.save(ARCHIVO_EXCEL)

        if reordenar_y_guardar_excel():
            messagebox.showinfo("Éxito", "Registro guardado y ordenado correctamente.")
            limpiar_campos()

    except PermissionError:
        messagebox.showerror(
            "Error de Permisos", "Cierre el archivo de Excel para poder guardar."
        )
    except Exception as e:
        messagebox.showerror("Error al Guardar", f"No se pudo guardar el registro: {e}")


def actualizar_registro():
    """Actualiza una fila existente en el archivo Excel."""
    global FILA_EDITANDO
    if not FILA_EDITANDO:
        return

    if not ARCHIVO_EXCEL:
        messagebox.showerror("Error", "No hay archivo seleccionado")
        return

    datos_actualizados = obtener_datos_formulario()
    if not datos_actualizados:
        return

    try:
        wb = load_workbook(ARCHIVO_EXCEL)
        ws = wb.active

        if ws is None:
            messagebox.showerror("Error", "No se pudo acceder a la hoja de trabajo")
            return

        ws.cell(row=FILA_EDITANDO, column=2, value=datos_actualizados["radicado"])
        ws.cell(row=FILA_EDITANDO, column=3, value=datos_actualizados["tipo"])
        ws.cell(row=FILA_EDITANDO, column=4, value=datos_actualizados["fecha"])
        ws.cell(row=FILA_EDITANDO, column=5, value=datos_actualizados["hora"])
        ws.cell(row=FILA_EDITANDO, column=6, value=datos_actualizados["juzgado"])
        ws.cell(row=FILA_EDITANDO, column=7, value=datos_actualizados["realizada_si"])
        ws.cell(row=FILA_EDITANDO, column=8, value=datos_actualizados["realizada_no"])
        for i, motivo in enumerate(datos_actualizados["motivos"]):
            ws.cell(row=FILA_EDITANDO, column=9 + i, value=motivo)
        ws.cell(row=FILA_EDITANDO, column=17, value=datos_actualizados["observaciones"])

        wb.save(ARCHIVO_EXCEL)

        if reordenar_y_guardar_excel():
            messagebox.showinfo("Éxito", "Registro actualizado correctamente.")
            desactivar_modo_edicion()

    except PermissionError:
        messagebox.showerror(
            "Error de Permisos", "Cierre el archivo de Excel para poder actualizar."
        )
    except Exception as e:
        messagebox.showerror(
            "Error al Actualizar", f"No se pudo actualizar el registro: {e}"
        )


def seleccionar_registro_para_editar():
    """Abre una ventana para que el usuario elija qué registro editar."""
    if not ARCHIVO_EXCEL:
        messagebox.showwarning("Sin Archivo", "Primero debe seleccionar un archivo.")
        return

    try:
        wb = load_workbook(ARCHIVO_EXCEL, read_only=True)
        ws = wb.active
        if ws is None:
            messagebox.showerror("Error", "No se pudo acceder a la hoja de trabajo")
            return
        registros = []
        for fila_num in range(FILA_INICIO_DATOS, ws.max_row + 1):
            datos_fila = [cell.value for cell in ws[fila_num]]
            if datos_fila[1]:  # Si hay radicado
                registros.append((fila_num, datos_fila))
    except Exception as e:
        messagebox.showerror(
            "Error de Lectura", f"No se pudieron leer los registros: {e}"
        )
        return

    if not registros:
        messagebox.showinfo("Vacío", "No hay registros para editar en el archivo.")
        return

    # Crear ventana de selección
    win_select = tk.Toplevel(ventana)
    win_select.title("Seleccionar Registro para Editar")
    win_select.geometry("800x400")
    win_select.grab_set()

    cols = ("N°", "Radicado", "Tipo", "Fecha", "Hora", "Juzgado")
    tree = ttk.Treeview(win_select, columns=cols, show="headings")
    for col in cols:
        tree.heading(col, text=col)
        tree.column(col, width=120, anchor="w")

    for fila_num, datos in registros:
        tree.insert(
            "",
            "end",
            values=(datos[0], datos[1], datos[2], datos[3], datos[4], datos[5]),
            tags=(fila_num,),
        )

    tree.pack(expand=True, fill="both", padx=10, pady=10)

    def on_edit():
        seleccion = tree.selection()
        if not seleccion:
            messagebox.showwarning(
                "Sin selección", "Por favor, seleccione un registro de la lista."
            )
            return

        fila_num_str = tree.item(seleccion[0])["tags"][0]
        global FILA_EDITANDO
        FILA_EDITANDO = int(fila_num_str)

        # Encontrar los datos completos de la fila seleccionada
        datos_completos = next(
            (datos for num, datos in registros if num == FILA_EDITANDO), None
        )

        if datos_completos:
            cargar_datos_para_edicion(datos_completos)
            activar_modo_edicion()
            win_select.destroy()

    btn_frame = tk.Frame(win_select)
    btn_frame.pack(pady=10)
    tk.Button(
        btn_frame, text="Editar Seleccionado", command=on_edit, bg="#28a745", fg="white"
    ).pack(side="left", padx=5)
    tk.Button(btn_frame, text="Cancelar", command=win_select.destroy).pack(
        side="left", padx=5
    )


def cancelar_edicion():
    """Cancela el modo de edición y revierte cambios no guardados."""
    if messagebox.askyesno(
        "Cancelar Edición",
        "¿Seguro que desea cancelar? Se perderán los cambios no guardados.",
    ):
        desactivar_modo_edicion()


def sumar_y_actualizar_totales(worksheet):
    """Calcula los totales de la hoja y los devuelve."""
    ws = worksheet
    totales_motivos = [0] * 8
    total_si = 0
    num_registros = 0

    for fila in range(FILA_INICIO_DATOS, FILA_MAXIMA_DATOS_PARA_LIMPIAR + 1):
        if ws.cell(row=fila, column=2).value:  # Si hay radicado
            num_registros += 1
            if ws.cell(row=fila, column=7).value == "SI":
                total_si += 1
            for i in range(8):
                if ws.cell(row=fila, column=9 + i).value:
                    totales_motivos[i] += 1
    return num_registros, total_si, totales_motivos


def actualizar_contador_registros():
    """Actualiza la etiqueta del contador de registros en la GUI."""
    if not ARCHIVO_EXCEL:
        if contador_registros_label:
            contador_registros_label.config(text="Registros: 0")
        return

    try:
        wb = load_workbook(ARCHIVO_EXCEL, read_only=True)
        ws = wb.active
        num_registros, _, _ = sumar_y_actualizar_totales(ws)
        if contador_registros_label:
            contador_registros_label.config(text=f"Registros: {num_registros}")
    except Exception:
        if contador_registros_label:
            contador_registros_label.config(text="Registros: ?")


# === Gestión de Archivos (Crear, Seleccionar, etc.) ===
def crear_nueva_copia():
    nombre = simpledialog.askstring(
        "Nuevo Archivo", "Nombre para la copia (ej: mayo_2025):"
    )
    if nombre:
        if not nombre.lower().endswith(".xlsx"):
            nombre += ".xlsx"
        try:
            crear_copia_plantilla(nombre)
            messagebox.showinfo("Éxito", f"Archivo '{nombre}' creado.")
        except FileExistsError:
            messagebox.showerror("Error", "Ya existe un archivo con ese nombre.")


archivo_actual_var = tk.StringVar(value="Ningún archivo seleccionado")
label_archivo_actual = tk.Label(
    ventana,
    textvariable=archivo_actual_var,
    bg="#f7f7fa",
    font=("Segoe UI", 9, "italic"),
    fg="#555",
)
label_archivo_actual.grid(
    row=11, column=0, columnspan=2, sticky="ew", padx=16, pady=(0, 12)
)


def seleccionar_archivo_trabajo():
    archivos = listar_archivos_creados()
    if not archivos:
        messagebox.showinfo("Sin Archivos", "No hay archivos de plantilla creados.")
        return

    seleccion = seleccionar_de_lista("Seleccionar Archivo de Trabajo", archivos)
    if seleccion:
        global ARCHIVO_EXCEL
        ARCHIVO_EXCEL = seleccionar_archivo(seleccion)
        archivo_actual_var.set(f"Trabajando con: {seleccion}")
        actualizar_contador_registros()
        messagebox.showinfo(
            "Archivo Seleccionado", f"Ahora trabajando con: {seleccion}"
        )


def eliminar_archivo_trabajo():
    archivos = listar_archivos_creados()
    if not archivos:
        return
    seleccion = seleccionar_de_lista("Eliminar Archivo", archivos)
    if seleccion and messagebox.askyesno(
        "Confirmar", f"¿Seguro que quiere eliminar '{seleccion}'?"
    ):
        eliminar_archivo(seleccion)
        messagebox.showinfo("Eliminado", f"Archivo '{seleccion}' eliminado.")
        if ARCHIVO_EXCEL and os.path.basename(ARCHIVO_EXCEL) == seleccion:
            global ARCHIVO_EXCEL
            ARCHIVO_EXCEL = None
            archivo_actual_var.set("Ningún archivo seleccionado")
            actualizar_contador_registros()


def descargar_archivo_trabajo():
    archivos = listar_archivos_creados()
    if not archivos:
        return
    seleccion = seleccionar_de_lista("Descargar Archivo", archivos)
    if seleccion:
        destino = descargar_archivo(seleccion)
        if destino:
            messagebox.showinfo("Descargado", f"Archivo guardado en:\n{destino}")


def seleccionar_de_lista(titulo, archivos):
    """Ventana genérica para seleccionar un ítem de una lista."""
    seleccion = tk.StringVar()
    win = tk.Toplevel(ventana)
    win.title(titulo)
    win.grab_set()
    tk.Label(win, text="Seleccione un archivo:").pack(padx=20, pady=10)
    combo = ttk.Combobox(win, values=archivos, state="readonly", width=40)
    combo.pack(padx=20, pady=5)
    if archivos:
        combo.current(0)

    def on_ok():
        seleccion.set(combo.get())
        win.destroy()

    tk.Button(win, text="Aceptar", command=on_ok).pack(pady=10)
    win.wait_window()
    return seleccion.get()


# Frame para botones de gestión de archivos
frame_gestion_archivos = tk.Frame(ventana, bg="#f7f7fa")
frame_gestion_archivos.grid(
    row=10, column=0, columnspan=2, pady=12, padx=16, sticky="ew"
)
tk.Button(frame_gestion_archivos, text="Crear Copia", command=crear_nueva_copia).pack(
    side="left", padx=4, expand=True, fill="x"
)
tk.Button(
    frame_gestion_archivos,
    text="Seleccionar Archivo",
    command=seleccionar_archivo_trabajo,
).pack(side="left", padx=4, expand=True, fill="x")
tk.Button(
    frame_gestion_archivos,
    text="Editar Registro",
    command=seleccionar_registro_para_editar,
    bg="#ffc107",
    fg="black",
).pack(side="left", padx=4, expand=True, fill="x")
tk.Button(
    frame_gestion_archivos, text="Eliminar Archivo", command=eliminar_archivo_trabajo
).pack(side="left", padx=4, expand=True, fill="x")
tk.Button(
    frame_gestion_archivos, text="Descargar Copia", command=descargar_archivo_trabajo
).pack(side="left", padx=4, expand=True, fill="x")


# === Configuración Final y Menú Contextual ===
def agregar_menu_contextual(widget):
    menu = tk.Menu(widget, tearoff=0)
    menu.add_command(label="Cortar", command=lambda: widget.event_generate("<<Cut>>"))
    menu.add_command(label="Copiar", command=lambda: widget.event_generate("<<Copy>>"))
    menu.add_command(label="Pegar", command=lambda: widget.event_generate("<<Paste>>"))
    widget.bind("<Button-3>", lambda e: menu.tk_popup(e.x_root, e.y_root))


# Aplicar menú contextual y validaciones
for widget in [
    entrada_radicado,
    entrada_tipo_tra,
    entrada_juzgado,
    entrada_observaciones,
]:
    agregar_menu_contextual(widget)
vcmd_hora = (ventana.register(lambda char: char.isdigit()), "%S")
spin_hora.config(validate="key", validatecommand=vcmd_hora)
spin_minuto.config(validate="key", validatecommand=vcmd_hora)

# Asignación de comandos a botones y eventos
btn_guardar.config(command=guardar_datos)
btn_actualizar.config(command=actualizar_registro)
btn_cancelar_edicion.config(command=cancelar_edicion)
combo_tipo.bind("<<ComboboxSelected>>", mostrar_entrada_otra)
combo_realizada.bind("<<ComboboxSelected>>", habilitar_motivos)

# Inicialización
limpiar_campos()
habilitar_motivos()

# Etiqueta de Crédito
label_credito = tk.Label(
    ventana,
    text="© 2025 - Hecho por Jose David Bustamante Sánchez",
    bg="#f7f7fa",
    font=("Segoe UI", 9, "italic"),
)
label_credito.grid(row=12, column=0, columnspan=2, pady=(0, 16))

ventana.mainloop()
